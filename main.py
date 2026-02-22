from datetime import date, datetime, timedelta, timezone
from io import StringIO
from typing import List, Optional
import csv
import os
import re
import shutil
import uuid
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import bcrypt
import qrcode
from PIL import Image, ImageDraw, ImageFont
from barcode import Code128
from barcode.writer import ImageWriter
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from jose import JWTError, jwt
from pydantic import BaseModel, ConfigDict, Field, field_validator
from sqlalchemy import (
    Boolean,
    Column,
    Date,
    DateTime,
    Float,
    ForeignKey,
    Integer,
    String,
    Text,
    create_engine,
    func,
    inspect,
    or_,
    text,
)
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import Session, relationship, sessionmaker

SECRET_KEY = os.getenv("ITAM_SECRET_KEY", "change-this-in-env")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
REFRESH_TOKEN_EXPIRE_DAYS = 7
LOGIN_LOCK_MINUTES = 15
MAX_FAILED_LOGINS = 5

DATABASE_URL = os.getenv(
    "ITAM_DB_URL",
    "mssql+pyodbc://apitam_user:MySqlPass123!@localhost\\SQLEXPRESS/itam_db?driver=ODBC+Driver+17+for+SQL+Server",
)

ENGINE = create_engine(DATABASE_URL, echo=False)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=ENGINE)
Base = declarative_base()

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/token")
oauth2_scheme_optional = OAuth2PasswordBearer(tokenUrl="/auth/token", auto_error=False)

PASSWORD_REGEX = re.compile(r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z\d]).{8,64}$")
EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
DEFAULT_ASSET_STATUSES = ["InStock", "Assigned", "UnderRepair", "Retired", "EndOfLife", "Scrapped", "Lost"]
try:
    IST = ZoneInfo("Asia/Kolkata")
except ZoneInfoNotFoundError:
    IST = timezone(timedelta(hours=5, minutes=30))


class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String(120), unique=True, index=True, nullable=False)
    email = Column(String(255), unique=True, index=True, nullable=False)
    hashed_password = Column(String(255), nullable=False)
    role = Column(String(50), nullable=False)
    is_active = Column(Boolean, default=True, nullable=False)

    assignments = relationship("Assignment", back_populates="assignee_user", foreign_keys="Assignment.user_id")
    assigned_actions = relationship("Assignment", back_populates="assigned_by_user", foreign_keys="Assignment.assigned_by")
    security = relationship("UserSecurity", back_populates="user", uselist=False)


class UserSecurity(Base):
    __tablename__ = "user_security"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), unique=True, nullable=False)
    must_change_password = Column(Boolean, default=False, nullable=False)
    failed_login_attempts = Column(Integer, default=0, nullable=False)
    locked_until = Column(DateTime, nullable=True)
    locked_until_utc = Column(DateTime, nullable=True)
    locked_until_local = Column(DateTime, nullable=True)
    last_login_at = Column(DateTime, nullable=True)
    last_login_at_utc = Column(DateTime, nullable=True)
    last_login_at_local = Column(DateTime, nullable=True)

    user = relationship("User", back_populates="security")


class Department(Base):
    __tablename__ = "departments"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(120), unique=True, nullable=False, index=True)
    is_active = Column(Boolean, default=True, nullable=False)


class Location(Base):
    __tablename__ = "locations"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(120), unique=True, nullable=False, index=True)
    is_active = Column(Boolean, default=True, nullable=False)


class Manufacturer(Base):
    __tablename__ = "manufacturers"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(120), unique=True, nullable=False, index=True)
    is_active = Column(Boolean, default=True, nullable=False)


class Vendor(Base):
    __tablename__ = "vendors"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(120), unique=True, nullable=False, index=True)
    is_active = Column(Boolean, default=True, nullable=False)


class AssetType(Base):
    __tablename__ = "asset_types"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(120), unique=True, nullable=False, index=True)
    mandatory_fields_csv = Column(String(500), nullable=True)
    default_values_json = Column(Text, nullable=True)
    is_active = Column(Boolean, default=True, nullable=False)


class StatusMaster(Base):
    __tablename__ = "status_master"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(120), unique=True, nullable=False, index=True)
    is_active = Column(Boolean, default=True, nullable=False)


class Asset(Base):
    __tablename__ = "assets"
    id = Column(Integer, primary_key=True, index=True)
    asset_id = Column(String(32), unique=True, index=True, nullable=False)
    asset_type = Column(String(120), nullable=False)
    serial_number = Column(String(120), index=True, nullable=False)
    manufacturer = Column(String(120), nullable=False)
    model = Column(String(120), nullable=False)
    purchase_date = Column(Date, nullable=False)
    warranty_start = Column(Date, nullable=False)
    warranty_end = Column(Date, nullable=False)
    vendor = Column(String(120), nullable=False)
    cost = Column(Float, nullable=True)
    location = Column(String(120), nullable=False)
    department = Column(String(120), nullable=False)
    status = Column(String(60), nullable=False, default="InStock")
    barcode_path = Column(String(255), nullable=True)

    assignments = relationship("Assignment", back_populates="asset")


class AssetHolder(Base):
    __tablename__ = "asset_holders"
    id = Column(Integer, primary_key=True, index=True)
    asset_id = Column(Integer, ForeignKey("assets.id"), unique=True, nullable=False, index=True)
    user_name = Column(String(120), nullable=True, index=True)
    email = Column(String(255), nullable=True, index=True)
    phone = Column(String(40), nullable=True, index=True)
    emp_id = Column(String(80), nullable=True, index=True)
    is_active = Column(Boolean, default=True, nullable=False)
    updated_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    updated_at_utc = Column(DateTime, nullable=True)
    updated_at_local = Column(DateTime, nullable=True)


class Assignment(Base):
    __tablename__ = "assignments"
    id = Column(Integer, primary_key=True, index=True)
    asset_id = Column(Integer, ForeignKey("assets.id"), nullable=False)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    department = Column(String(120), nullable=True)
    location = Column(String(120), nullable=True)
    assign_date = Column(Date, nullable=False)
    expected_return = Column(Date, nullable=True)
    remarks = Column(String(500), nullable=False)
    assigned_by = Column(Integer, ForeignKey("users.id"), nullable=False)

    asset = relationship("Asset", back_populates="assignments")
    assignee_user = relationship("User", back_populates="assignments", foreign_keys=[user_id])
    assigned_by_user = relationship("User", back_populates="assigned_actions", foreign_keys=[assigned_by])


class AssignmentEvent(Base):
    __tablename__ = "assignment_events"
    id = Column(Integer, primary_key=True, index=True)
    asset_id = Column(Integer, ForeignKey("assets.id"), nullable=False, index=True)
    event_type = Column(String(30), nullable=False)
    from_user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    to_user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    from_department = Column(String(120), nullable=True)
    to_department = Column(String(120), nullable=True)
    from_location = Column(String(120), nullable=True)
    to_location = Column(String(120), nullable=True)
    expected_return = Column(Date, nullable=True)
    remarks = Column(String(500), nullable=False)
    created_by = Column(Integer, ForeignKey("users.id"), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    created_at_utc = Column(DateTime, nullable=True)
    created_at_local = Column(DateTime, nullable=True)


class RefreshToken(Base):
    __tablename__ = "refresh_tokens"
    id = Column(Integer, primary_key=True, index=True)
    jti = Column(String(64), unique=True, index=True, nullable=False)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    expires_at = Column(DateTime, nullable=False)
    expires_at_utc = Column(DateTime, nullable=True)
    expires_at_local = Column(DateTime, nullable=True)
    revoked = Column(Boolean, default=False, nullable=False)


class AuditLog(Base):
    __tablename__ = "audit_logs"
    id = Column(Integer, primary_key=True, index=True)
    actor_user_id = Column(Integer, ForeignKey("users.id"), nullable=True)
    action = Column(String(120), nullable=False)
    entity_type = Column(String(120), nullable=False)
    entity_id = Column(String(120), nullable=False)
    details = Column(String(1000), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    created_at_utc = Column(DateTime, nullable=True)
    created_at_local = Column(DateTime, nullable=True)


Base.metadata.create_all(bind=ENGINE)


class TokenOut(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str
    must_change_password: bool


class RefreshIn(BaseModel):
    refresh_token: str


class PasswordChangeIn(BaseModel):
    old_password: str
    new_password: str

    @field_validator("new_password")
    @classmethod
    def validate_new_password(cls, value: str) -> str:
        if not PASSWORD_REGEX.match(value):
            raise ValueError("Password must be 8-64 chars with upper, lower, digit, and special character")
        return value


class UserCreate(BaseModel):
    username: str = Field(min_length=3, max_length=120)
    email: str
    password: str
    role: str = Field(pattern="^(Admin|ITUser|Viewer)$")

    @field_validator("password")
    @classmethod
    def validate_password(cls, value: str) -> str:
        if not PASSWORD_REGEX.match(value):
            raise ValueError("Password must be 8-64 chars with upper, lower, digit, and special character")
        return value

    @field_validator("email")
    @classmethod
    def validate_email(cls, value: str) -> str:
        if not EMAIL_REGEX.match(value):
            raise ValueError("Invalid email format")
        return value


class UserOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    username: str
    email: str
    role: str
    is_active: bool


class MeOut(BaseModel):
    id: int
    username: str
    email: str
    role: str
    is_active: bool
    must_change_password: bool


class UserStatusUpdate(BaseModel):
    is_active: bool


class PasswordResetIn(BaseModel):
    temporary_password: str

    @field_validator("temporary_password")
    @classmethod
    def validate_temp_password(cls, value: str) -> str:
        if not PASSWORD_REGEX.match(value):
            raise ValueError("Password must be 8-64 chars with upper, lower, digit, and special character")
        return value


class UserActivityOut(BaseModel):
    user_id: int
    username: str
    last_login_at: Optional[datetime]
    failed_login_attempts: int
    locked_until: Optional[datetime]


class DepartmentIn(BaseModel):
    name: str = Field(min_length=2, max_length=120)


class DepartmentOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    is_active: bool


class LocationIn(BaseModel):
    name: str = Field(min_length=2, max_length=120)


class LocationOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    is_active: bool


class ManufacturerIn(BaseModel):
    name: str = Field(min_length=2, max_length=120)


class ManufacturerOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    is_active: bool


class VendorIn(BaseModel):
    name: str = Field(min_length=2, max_length=120)


class VendorOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    is_active: bool


class AssetTypeIn(BaseModel):
    name: str = Field(min_length=2, max_length=120)
    mandatory_fields_csv: Optional[str] = None
    default_values_json: Optional[str] = None


class AssetTypeOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    mandatory_fields_csv: Optional[str]
    default_values_json: Optional[str]
    is_active: bool


class StatusIn(BaseModel):
    name: str = Field(min_length=2, max_length=120)


class StatusOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    is_active: bool


class AssetCreate(BaseModel):
    asset_type: str
    serial_number: str
    manufacturer: str
    model: str
    purchase_date: date
    warranty_start: date
    warranty_end: date
    vendor: str
    cost: Optional[float] = None
    location: str
    department: str
    status: str = "InStock"


class AssetUpdate(BaseModel):
    asset_type: Optional[str] = None
    serial_number: Optional[str] = None
    manufacturer: Optional[str] = None
    model: Optional[str] = None
    purchase_date: Optional[date] = None
    warranty_start: Optional[date] = None
    warranty_end: Optional[date] = None
    vendor: Optional[str] = None
    cost: Optional[float] = None
    location: Optional[str] = None
    department: Optional[str] = None
    status: Optional[str] = None


class AssetOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    asset_id: str
    asset_type: str
    serial_number: str
    manufacturer: str
    model: str
    purchase_date: date
    warranty_start: date
    warranty_end: date
    vendor: str
    cost: Optional[float]
    location: str
    department: str
    status: str
    barcode_path: Optional[str]


class AssignmentCreate(BaseModel):
    asset_id: int
    user_id: Optional[int] = None
    assignee_name: Optional[str] = None
    assignee_email: Optional[str] = None
    assignee_phone: Optional[str] = None
    assignee_emp_id: Optional[str] = None
    department: Optional[str] = None
    location: Optional[str] = None
    expected_return: Optional[date] = None
    remarks: str = Field(min_length=2, max_length=500)


class AssignmentOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    asset_id: int
    user_id: Optional[int]
    department: Optional[str]
    location: Optional[str]
    assign_date: date
    expected_return: Optional[date]
    remarks: str
    assigned_by: int


class AssignmentEventOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    asset_id: int
    event_type: str
    from_user_id: Optional[int]
    to_user_id: Optional[int]
    from_department: Optional[str]
    to_department: Optional[str]
    from_location: Optional[str]
    to_location: Optional[str]
    expected_return: Optional[date]
    remarks: str
    created_by: int
    created_at: datetime


class DashboardSummaryOut(BaseModel):
    total_assets: int
    assigned_assets: int
    unassigned_assets: int
    under_repair_assets: int
    warranty_expiring_30_days: int
    by_type: List[dict]
    by_department: List[dict]


class WarrantyAlertOut(BaseModel):
    asset_id: str
    serial_number: str
    warranty_end: date
    days_left: int


class AuditLogOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    actor_user_id: Optional[int]
    action: str
    entity_type: str
    entity_id: str
    details: Optional[str]
    created_at: datetime


class TimelineEventOut(BaseModel):
    timestamp: datetime
    event_type: str
    details: str


class AssetTableRowOut(BaseModel):
    id: int
    asset_id: str
    asset_type: str
    serial_number: str
    manufacturer: str
    model: str
    status: str
    location: str
    department: str
    assignee_name: Optional[str] = None
    assignee_email: Optional[str] = None
    assignee_phone: Optional[str] = None
    assignee_emp_id: Optional[str] = None


class AssigneeSummaryOut(BaseModel):
    user_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    emp_id: Optional[str] = None
    current_assets: List[dict]
    historical_assets: List[dict]


class AssetLifecycleIn(BaseModel):
    action: str = Field(pattern="^(scrap|end_of_life|lost)$")
    remarks: str = Field(min_length=2, max_length=500)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def utc_now_naive() -> datetime:
    return datetime.utcnow()


def ist_from_utc_naive(dt_utc: datetime) -> datetime:
    return dt_utc.replace(tzinfo=timezone.utc).astimezone(IST).replace(tzinfo=None)


def now_utc_local_pair() -> tuple[datetime, datetime]:
    utc_dt = utc_now_naive()
    return utc_dt, ist_from_utc_naive(utc_dt)


def as_ist_aware(dt_local: Optional[datetime], fallback_utc: Optional[datetime] = None) -> Optional[datetime]:
    if dt_local:
        return dt_local.replace(tzinfo=IST)
    if fallback_utc:
        return fallback_utc.replace(tzinfo=timezone.utc).astimezone(IST)
    return None


def ensure_user_security(db: Session, user: User) -> UserSecurity:
    sec = db.query(UserSecurity).filter(UserSecurity.user_id == user.id).first()
    if not sec:
        sec = UserSecurity(user_id=user.id)
        db.add(sec)
        db.commit()
        db.refresh(sec)
    return sec


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        return bcrypt.checkpw(plain_password.encode("utf-8"), hashed_password.encode("utf-8"))
    except Exception:
        return False


def create_token(sub: str, token_type: str, expires_delta: timedelta, jti: Optional[str] = None) -> str:
    now = datetime.now(timezone.utc)
    payload = {
        "sub": sub,
        "type": token_type,
        "iat": int(now.timestamp()),
        "exp": int((now + expires_delta).timestamp()),
    }
    if jti:
        payload["jti"] = jti
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def create_access_token(username: str) -> str:
    return create_token(sub=username, token_type="access", expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))


def create_refresh_token(username: str) -> tuple[str, str, datetime, datetime]:
    jti = str(uuid.uuid4())
    expires_at = datetime.utcnow() + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    expires_at_local = ist_from_utc_naive(expires_at)
    token = create_token(sub=username, token_type="refresh", expires_delta=timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS), jti=jti)
    return token, jti, expires_at, expires_at_local


def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Could not validate credentials")


def log_audit(db: Session, action: str, entity_type: str, entity_id: str, actor_user_id: Optional[int] = None, details: Optional[str] = None) -> None:
    now_utc, now_local = now_utc_local_pair()
    entry = AuditLog(
        actor_user_id=actor_user_id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
        created_at=now_utc,
        created_at_utc=now_utc,
        created_at_local=now_local,
    )
    db.add(entry)
    db.commit()


def require_active_status(status_name: str, db: Session) -> None:
    if db.query(StatusMaster.id).count() == 0:
        for status_name_seed in DEFAULT_ASSET_STATUSES:
            db.add(StatusMaster(name=status_name_seed, is_active=True))
        db.commit()
    status_row = db.query(StatusMaster).filter(StatusMaster.name == status_name, StatusMaster.is_active == True).first()
    if not status_row:
        raise HTTPException(status_code=400, detail=f"Invalid status: {status_name}")


def upsert_asset_holder(
    db: Session,
    asset_db_id: int,
    assignee_name: Optional[str],
    assignee_email: Optional[str],
    assignee_phone: Optional[str],
    assignee_emp_id: Optional[str],
    is_active: bool,
) -> None:
    row = db.query(AssetHolder).filter(AssetHolder.asset_id == asset_db_id).first()
    if not row:
        row = AssetHolder(asset_id=asset_db_id)
        db.add(row)
    row.user_name = assignee_name
    row.email = assignee_email
    row.phone = assignee_phone
    row.emp_id = assignee_emp_id
    row.is_active = is_active
    now_utc, now_local = now_utc_local_pair()
    row.updated_at = now_utc
    row.updated_at_utc = now_utc
    row.updated_at_local = now_local


def normalize_code_token(value: Optional[str], fallback: str, max_len: int = 3) -> str:
    token = re.sub(r"[^A-Za-z0-9]", "", (value or "").upper())
    if not token:
        token = fallback
    return token[:max_len]


def generate_asset_code(payload: "AssetCreate", db: Session) -> str:
    # Format: ITAM-<TYPE>-<LOC>-<YEAR>-<SEQ>, example: ITAM-LAP-HQ-2026-00001
    year = datetime.utcnow().year
    type_code = normalize_code_token(payload.asset_type, "GEN", 3)
    location_code = normalize_code_token(payload.location, "LOC", 3)
    prefix = f"ITAM-{type_code}-{location_code}-{year}"

    last_row = (
        db.query(Asset.asset_id)
        .filter(Asset.asset_id.like(f"{prefix}-%"))
        .order_by(Asset.asset_id.desc())
        .first()
    )

    next_seq = 1
    if last_row and last_row[0]:
        parts = last_row[0].split("-")
        if parts:
            try:
                next_seq = int(parts[-1]) + 1
            except ValueError:
                next_seq = 1

    while True:
        candidate = f"{prefix}-{next_seq:05d}"
        exists = db.query(Asset.id).filter(Asset.asset_id == candidate).first()
        if not exists:
            return candidate
        next_seq += 1


def normalize_row_key(key: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (key or "").strip().lower()).strip("_")


def parse_upload_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value).strip()
    if not raw:
        return ""

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Invalid date '{raw}'. Use YYYY-MM-DD")


def parse_upload_row_to_asset_payload(row: dict) -> AssetCreate:
    norm = {normalize_row_key(str(k)): v for k, v in row.items()}

    def val(key: str, default: str = ""):
        v = norm.get(key, default)
        if v is None:
            return ""
        return str(v).strip()

    cost_raw = norm.get("cost", None)
    cost_val = None
    if cost_raw not in [None, ""]:
        try:
            cost_val = float(str(cost_raw).strip())
        except ValueError:
            raise ValueError(f"Invalid cost '{cost_raw}'")

    return AssetCreate(
        asset_type=val("asset_type"),
        serial_number=val("serial_number"),
        manufacturer=val("manufacturer"),
        model=val("model"),
        purchase_date=parse_upload_date(norm.get("purchase_date")),
        warranty_start=parse_upload_date(norm.get("warranty_start")),
        warranty_end=parse_upload_date(norm.get("warranty_end")),
        vendor=val("vendor"),
        cost=cost_val,
        location=val("location"),
        department=val("department"),
        status=val("status", "InStock") or "InStock",
    )


def create_asset_row(payload: AssetCreate, db: Session) -> Asset:
    require_active_status(payload.status, db)
    existing_sn = db.query(Asset).filter(Asset.serial_number == payload.serial_number).first()
    if existing_sn:
        raise HTTPException(status_code=400, detail=f"Serial number already exists: {payload.serial_number}")
    asset_code = generate_asset_code(payload, db)
    asset = Asset(**payload.model_dump(), asset_id=asset_code)
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)) -> User:
    payload = decode_token(token)
    if payload.get("type") != "access":
        raise HTTPException(status_code=401, detail="Invalid token type")
    username = payload.get("sub")
    if not username:
        raise HTTPException(status_code=401, detail="Invalid token payload")
    user = db.query(User).filter(User.username == username).first()
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="User inactive or not found")
    return user


def get_current_user_optional(token: Optional[str] = Depends(oauth2_scheme_optional), db: Session = Depends(get_db)) -> Optional[User]:
    if not token:
        return None
    payload = decode_token(token)
    if payload.get("type") != "access":
        raise HTTPException(status_code=401, detail="Invalid token type")
    username = payload.get("sub")
    if not username:
        raise HTTPException(status_code=401, detail="Invalid token payload")
    user = db.query(User).filter(User.username == username).first()
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="User inactive or not found")
    return user


def require_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role != "Admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user


def require_it_or_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role not in ["Admin", "ITUser"]:
        raise HTTPException(status_code=403, detail="IT/Admin access required")
    return current_user


def asset_qr_path(asset_id: str) -> str:
    return f"static/barcodes/{asset_id}_qr.png"


def generate_qr_label(asset: Asset) -> str:
    # QR payload is the asset_id so scanning shows the unique tag directly.
    qr = qrcode.QRCode(version=2, box_size=8, border=2)
    qr.add_data(asset.asset_id)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    qr_img = qr_img.resize((240, 240))

    label = Image.new("RGB", (1000, 300), "white")
    label.paste(qr_img, (20, 20))

    draw = ImageDraw.Draw(label)
    try:
        title_font = ImageFont.truetype("arial.ttf", 42)
        text_font = ImageFont.truetype("arial.ttf", 28)
        tag_font = ImageFont.truetype("arial.ttf", 34)
    except Exception:
        title_font = ImageFont.load_default()
        text_font = ImageFont.load_default()
        tag_font = ImageFont.load_default()

    draw.text((290, 28), f"{asset.asset_type}", fill="black", font=title_font)
    draw.text((290, 92), "Initial Location", fill="black", font=text_font)
    draw.text((290, 124), f"{asset.location}", fill="black", font=tag_font)
    draw.text((290, 176), "Tag", fill="black", font=text_font)
    draw.text((290, 206), f"{asset.asset_id}", fill="black", font=tag_font)
    draw.text((72, 266), asset.asset_id, fill="black", font=tag_font)
    draw.text((560, 92), f"Serial: {asset.serial_number}", fill="black", font=text_font)

    out_path = asset_qr_path(asset.asset_id)
    label.save(out_path)
    return out_path


app = FastAPI(title="InfraTrack API", version="2.0.0")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
os.makedirs("static/barcodes", exist_ok=True)
os.makedirs("static/branding", exist_ok=True)

APP_NAME = "InfraTrack"


class BrandingOut(BaseModel):
    app_name: str
    logo_url: Optional[str] = None


def ensure_datetime_dual_columns() -> None:
    # Lightweight schema migration for existing databases.
    required_columns = {
        "user_security": ["locked_until_utc", "locked_until_local", "last_login_at_utc", "last_login_at_local"],
        "asset_holders": ["updated_at_utc", "updated_at_local"],
        "assignment_events": ["created_at_utc", "created_at_local"],
        "refresh_tokens": ["expires_at_utc", "expires_at_local"],
        "audit_logs": ["created_at_utc", "created_at_local"],
    }
    insp = inspect(ENGINE)
    with ENGINE.begin() as conn:
        for table_name, cols in required_columns.items():
            existing = {c["name"] for c in insp.get_columns(table_name)}
            for col in cols:
                if col not in existing:
                    conn.execute(text(f"ALTER TABLE {table_name} ADD {col} DATETIME NULL"))


def get_logo_file() -> Optional[str]:
    for name in os.listdir("static/branding"):
        lower = name.lower()
        if lower.startswith("company_logo.") and lower.split(".")[-1] in ["png", "jpg", "jpeg", "webp", "gif"]:
            return os.path.join("static/branding", name).replace("\\", "/")
    return None


@app.on_event("startup")
def seed_defaults():
    ensure_datetime_dual_columns()
    db = SessionLocal()
    try:
        for status_name in DEFAULT_ASSET_STATUSES:
            exists = db.query(StatusMaster).filter(StatusMaster.name == status_name).first()
            if not exists:
                db.add(StatusMaster(name=status_name, is_active=True))
        for row_model, values in [
            (AssetType, ["Laptop", "Desktop", "Monitor", "Printer", "Scanner", "Network Equipment", "Server", "Software License"]),
            (Department, ["IT", "Finance", "HR", "Operations"]),
            (Location, ["HQ", "Warehouse"]),
            (Manufacturer, ["Dell", "HP", "Lenovo"]),
            (Vendor, ["Default Vendor"]),
        ]:
            for value in values:
                exists = db.query(row_model).filter(row_model.name == value).first()
                if not exists:
                    db.add(row_model(name=value, is_active=True))
        db.commit()
    finally:
        db.close()

@app.post("/auth/token", response_model=TokenOut)
def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == form_data.username).first()
    if not user:
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="User disabled")

    sec = ensure_user_security(db, user)
    now = datetime.utcnow()
    if sec.locked_until and sec.locked_until > now:
        raise HTTPException(status_code=423, detail="User account is temporarily locked")

    if not verify_password(form_data.password, user.hashed_password):
        sec.failed_login_attempts += 1
        if sec.failed_login_attempts >= MAX_FAILED_LOGINS:
            lock_utc = now + timedelta(minutes=LOGIN_LOCK_MINUTES)
            sec.locked_until = lock_utc
            sec.locked_until_utc = lock_utc
            sec.locked_until_local = ist_from_utc_naive(lock_utc)
            sec.failed_login_attempts = 0
        db.commit()
        raise HTTPException(status_code=401, detail="Incorrect username or password")

    sec.failed_login_attempts = 0
    sec.locked_until = None
    sec.locked_until_utc = None
    sec.locked_until_local = None
    sec.last_login_at = now
    sec.last_login_at_utc = now
    sec.last_login_at_local = ist_from_utc_naive(now)
    db.commit()

    access_token = create_access_token(user.username)
    refresh_token, jti, expires_at, expires_at_local = create_refresh_token(user.username)
    db.add(
        RefreshToken(
            jti=jti,
            user_id=user.id,
            expires_at=expires_at,
            expires_at_utc=expires_at,
            expires_at_local=expires_at_local,
            revoked=False,
        )
    )
    db.commit()

    log_audit(db, action="LOGIN_SUCCESS", entity_type="User", entity_id=str(user.id), actor_user_id=user.id, details=f"username={user.username}")

    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "must_change_password": sec.must_change_password,
    }


@app.post("/auth/refresh", response_model=TokenOut)
def refresh_access_token(payload: RefreshIn, db: Session = Depends(get_db)):
    decoded = decode_token(payload.refresh_token)
    if decoded.get("type") != "refresh":
        raise HTTPException(status_code=401, detail="Invalid token type")
    username = decoded.get("sub")
    jti = decoded.get("jti")
    if not username or not jti:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

    token_row = db.query(RefreshToken).filter(RefreshToken.jti == jti).first()
    if (not token_row or token_row.revoked or token_row.expires_at < datetime.utcnow()):
        raise HTTPException(status_code=401, detail="Refresh token revoked or expired")

    user = db.query(User).filter(User.id == token_row.user_id).first()
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="User inactive or not found")

    sec = ensure_user_security(db, user)
    new_access = create_access_token(user.username)
    new_refresh, new_jti, expires_at, expires_at_local = create_refresh_token(user.username)

    token_row.revoked = True
    db.add(
        RefreshToken(
            jti=new_jti,
            user_id=user.id,
            expires_at=expires_at,
            expires_at_utc=expires_at,
            expires_at_local=expires_at_local,
            revoked=False,
        )
    )
    db.commit()

    return {
        "access_token": new_access,
        "refresh_token": new_refresh,
        "token_type": "bearer",
        "must_change_password": sec.must_change_password,
    }


@app.post("/auth/change-password")
def change_password(payload: PasswordChangeIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not verify_password(payload.old_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Old password is incorrect")
    current_user.hashed_password = hash_password(payload.new_password)
    sec = ensure_user_security(db, current_user)
    sec.must_change_password = False
    db.commit()

    log_audit(db, action="PASSWORD_CHANGED", entity_type="User", entity_id=str(current_user.id), actor_user_id=current_user.id)
    return {"message": "Password changed successfully"}


@app.get("/auth/me", response_model=MeOut)
def get_me(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    sec = ensure_user_security(db, current_user)
    return {
        "id": current_user.id,
        "username": current_user.username,
        "email": current_user.email,
        "role": current_user.role,
        "is_active": current_user.is_active,
        "must_change_password": sec.must_change_password,
    }


@app.post("/users/", response_model=UserOut)
def create_user(payload: UserCreate, current_user: Optional[User] = Depends(get_current_user_optional), db: Session = Depends(get_db)):
    users_exist = db.query(User.id).first() is not None
    if users_exist:
        if not current_user:
            raise HTTPException(status_code=401, detail="Authentication required")
        if current_user.role != "Admin":
            raise HTTPException(status_code=403, detail="Admin access required")

    exists = db.query(User).filter(or_(User.username == payload.username, User.email == payload.email)).first()
    if exists:
        raise HTTPException(status_code=400, detail="Username or email already exists")

    assigned_role = "Admin" if not users_exist else payload.role
    user = User(username=payload.username, email=payload.email, hashed_password=hash_password(payload.password), role=assigned_role, is_active=True)
    db.add(user)
    db.commit()
    db.refresh(user)

    sec = ensure_user_security(db, user)
    sec.must_change_password = users_exist
    db.commit()

    log_audit(
        db,
        action="USER_CREATED",
        entity_type="User",
        entity_id=str(user.id),
        actor_user_id=current_user.id if current_user else user.id,
        details=f"role={user.role}",
    )
    return user


@app.get("/users/", response_model=List[UserOut])
def list_users(_: User = Depends(require_admin), db: Session = Depends(get_db)):
    return db.query(User).order_by(User.id.asc()).all()


@app.delete("/users/{user_id}")
def delete_user(user_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    # Soft-delete to preserve audit/assignment history that references this user.
    old_username = user.username
    suffix = str(uuid.uuid4())[:8]
    user.is_active = False
    user.username = f"deleted_{suffix}"
    user.email = f"deleted_{suffix}@invalid.local"
    db.commit()
    log_audit(
        db,
        action="USER_DEACTIVATED",
        entity_type="User",
        entity_id=str(user_id),
        actor_user_id=current_user.id,
        details=f"old_username={old_username}",
    )
    return {"message": "User deactivated"}


@app.patch("/users/{user_id}/status", response_model=UserOut)
def update_user_status(user_id: int, payload: UserStatusUpdate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.is_active = payload.is_active
    db.commit()
    db.refresh(user)
    log_audit(db, action="USER_STATUS_CHANGED", entity_type="User", entity_id=str(user.id), actor_user_id=current_user.id, details=f"is_active={user.is_active}")
    return user


@app.post("/users/{user_id}/reset-password")
def reset_user_password(user_id: int, payload: PasswordResetIn, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.hashed_password = hash_password(payload.temporary_password)
    sec = ensure_user_security(db, user)
    sec.must_change_password = True
    sec.failed_login_attempts = 0
    sec.locked_until = None
    db.commit()
    log_audit(db, action="USER_PASSWORD_RESET", entity_type="User", entity_id=str(user.id), actor_user_id=current_user.id)
    return {"message": "Temporary password set. User must change password on next login."}


@app.get("/users/{user_id}/activity", response_model=UserActivityOut)
def get_user_activity(user_id: int, _: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    sec = ensure_user_security(db, user)
    return {
        "user_id": user.id,
        "username": user.username,
        "last_login_at": sec.last_login_at_local or sec.last_login_at,
        "failed_login_attempts": sec.failed_login_attempts,
        "locked_until": sec.locked_until_local or sec.locked_until,
    }


@app.post("/masters/departments", response_model=DepartmentOut)
def create_department(payload: DepartmentIn, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    existing = db.query(Department).filter(Department.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Department already exists")
    row = Department(name=payload.name.strip())
    db.add(row)
    db.commit()
    db.refresh(row)
    log_audit(db, action="DEPARTMENT_CREATED", entity_type="Department", entity_id=str(row.id), actor_user_id=current_user.id)
    return row


@app.get("/masters/departments", response_model=List[DepartmentOut])
def list_departments(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Department).order_by(Department.name.asc()).all()


@app.post("/masters/locations", response_model=LocationOut)
def create_location(payload: LocationIn, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    existing = db.query(Location).filter(Location.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Location already exists")
    row = Location(name=payload.name.strip())
    db.add(row)
    db.commit()
    db.refresh(row)
    log_audit(db, action="LOCATION_CREATED", entity_type="Location", entity_id=str(row.id), actor_user_id=current_user.id)
    return row


@app.get("/masters/locations", response_model=List[LocationOut])
def list_locations(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Location).order_by(Location.name.asc()).all()


@app.post("/masters/manufacturers", response_model=ManufacturerOut)
def create_manufacturer(payload: ManufacturerIn, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    existing = db.query(Manufacturer).filter(Manufacturer.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Manufacturer already exists")
    row = Manufacturer(name=payload.name.strip())
    db.add(row)
    db.commit()
    db.refresh(row)
    log_audit(db, action="MANUFACTURER_CREATED", entity_type="Manufacturer", entity_id=str(row.id), actor_user_id=current_user.id)
    return row


@app.get("/masters/manufacturers", response_model=List[ManufacturerOut])
def list_manufacturers(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Manufacturer).order_by(Manufacturer.name.asc()).all()


@app.post("/masters/vendors", response_model=VendorOut)
def create_vendor(payload: VendorIn, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    existing = db.query(Vendor).filter(Vendor.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Vendor already exists")
    row = Vendor(name=payload.name.strip())
    db.add(row)
    db.commit()
    db.refresh(row)
    log_audit(db, action="VENDOR_CREATED", entity_type="Vendor", entity_id=str(row.id), actor_user_id=current_user.id)
    return row


@app.get("/masters/vendors", response_model=List[VendorOut])
def list_vendors(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Vendor).order_by(Vendor.name.asc()).all()


@app.post("/masters/asset-types", response_model=AssetTypeOut)
def create_asset_type(payload: AssetTypeIn, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    existing = db.query(AssetType).filter(AssetType.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Asset type already exists")
    row = AssetType(name=payload.name.strip(), mandatory_fields_csv=payload.mandatory_fields_csv, default_values_json=payload.default_values_json)
    db.add(row)
    db.commit()
    db.refresh(row)
    log_audit(db, action="ASSET_TYPE_CREATED", entity_type="AssetType", entity_id=str(row.id), actor_user_id=current_user.id)
    return row


@app.get("/masters/asset-types", response_model=List[AssetTypeOut])
def list_asset_types(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(AssetType).order_by(AssetType.name.asc()).all()


@app.post("/masters/statuses", response_model=StatusOut)
def create_status(payload: StatusIn, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    existing = db.query(StatusMaster).filter(StatusMaster.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Status already exists")
    row = StatusMaster(name=payload.name.strip())
    db.add(row)
    db.commit()
    db.refresh(row)
    log_audit(db, action="STATUS_CREATED", entity_type="StatusMaster", entity_id=str(row.id), actor_user_id=current_user.id)
    return row


@app.get("/masters/statuses", response_model=List[StatusOut])
def list_statuses(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(StatusMaster).order_by(StatusMaster.name.asc()).all()


@app.post("/assets/", response_model=AssetOut)
def create_asset(payload: AssetCreate, current_user: User = Depends(require_it_or_admin), db: Session = Depends(get_db)):
    try:
        asset = create_asset_row(payload, db)
    except HTTPException:
        raise
    except SQLAlchemyError:
        db.rollback()
        raise HTTPException(status_code=500, detail="Database error while creating asset")

    try:
        barcode_file = f"static/barcodes/{asset.asset_id}.png"
        code = Code128(asset.asset_id, writer=ImageWriter())
        code.save(barcode_file[:-4])
        asset.barcode_path = barcode_file

        generate_qr_label(asset)
        db.commit()
        db.refresh(asset)
    except Exception:
        db.rollback()
        # Keep asset creation successful even if media generation fails.

    log_audit(db, action="ASSET_CREATED", entity_type="Asset", entity_id=str(asset.id), actor_user_id=current_user.id, details=f"asset_id={asset.asset_id}")
    return asset


@app.get("/assets/", response_model=List[AssetOut])
def list_assets(
    q: Optional[str] = None,
    status_filter: Optional[str] = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    query = db.query(Asset)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Asset.asset_id.ilike(like),
                Asset.serial_number.ilike(like),
                Asset.manufacturer.ilike(like),
                Asset.model.ilike(like),
            )
        )
    if status_filter:
        query = query.filter(Asset.status == status_filter)
    return query.order_by(Asset.id.desc()).limit(max(1, min(limit, 1000))).all()


@app.get("/assets/table", response_model=List[AssetTableRowOut])
def list_assets_table(
    q: Optional[str] = None,
    status_filter: Optional[str] = None,
    limit: int = 200,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    query = db.query(Asset, AssetHolder).outerjoin(AssetHolder, AssetHolder.asset_id == Asset.id)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Asset.asset_id.ilike(like),
                Asset.serial_number.ilike(like),
                Asset.manufacturer.ilike(like),
                Asset.model.ilike(like),
                AssetHolder.user_name.ilike(like),
                AssetHolder.email.ilike(like),
                AssetHolder.emp_id.ilike(like),
            )
        )
    if status_filter:
        query = query.filter(Asset.status == status_filter)

    rows = query.order_by(Asset.id.desc()).limit(max(1, min(limit, 1000))).all()
    out = []
    for asset, holder in rows:
        out.append(
            {
                "id": asset.id,
                "asset_id": asset.asset_id,
                "asset_type": asset.asset_type,
                "serial_number": asset.serial_number,
                "manufacturer": asset.manufacturer,
                "model": asset.model,
                "status": asset.status,
                "location": asset.location,
                "department": asset.department,
                "assignee_name": holder.user_name if holder and holder.is_active else None,
                "assignee_email": holder.email if holder and holder.is_active else None,
                "assignee_phone": holder.phone if holder and holder.is_active else None,
                "assignee_emp_id": holder.emp_id if holder and holder.is_active else None,
            }
        )
    return out


@app.post("/assets/bulk-upload")
async def bulk_upload_assets(
    file: UploadFile = File(...),
    current_user: User = Depends(require_it_or_admin),
    db: Session = Depends(get_db),
):
    filename = (file.filename or "").lower()
    created = 0
    failed = 0
    errors = []

    if filename.endswith(".csv"):
        content = (await file.read()).decode("utf-8-sig")
        reader = csv.DictReader(StringIO(content))
        for idx, row in enumerate(reader, start=2):
            try:
                payload = parse_upload_row_to_asset_payload(row)
                asset = create_asset_row(payload, db)
                created += 1
                log_audit(db, action="ASSET_CREATED_BULK", entity_type="Asset", entity_id=str(asset.id), actor_user_id=current_user.id, details=f"asset_id={asset.asset_id}")
            except Exception as exc:
                db.rollback()
                failed += 1
                errors.append({"line": idx, "error": str(exc)})
    elif filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise HTTPException(status_code=400, detail="XLSX upload requires openpyxl. Install with: pip install openpyxl")

        content = await file.read()
        from io import BytesIO
        wb = load_workbook(filename=BytesIO(content), data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        for ridx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = {header[i]: row_values[i] for i in range(min(len(header), len(row_values)))}
            try:
                payload = parse_upload_row_to_asset_payload(row)
                asset = create_asset_row(payload, db)
                created += 1
                log_audit(db, action="ASSET_CREATED_BULK", entity_type="Asset", entity_id=str(asset.id), actor_user_id=current_user.id, details=f"asset_id={asset.asset_id}")
            except Exception as exc:
                db.rollback()
                failed += 1
                errors.append({"line": ridx, "error": str(exc)})
    else:
        raise HTTPException(status_code=400, detail="Only CSV and XLSX uploads are supported")

    return {"created": created, "failed": failed, "errors": errors[:50]}


@app.get("/assets/sample-template.csv")
def download_asset_upload_template(_: User = Depends(get_current_user)):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "asset_type",
            "serial_number",
            "manufacturer",
            "model",
            "purchase_date",
            "warranty_start",
            "warranty_end",
            "vendor",
            "cost",
            "location",
            "department",
            "status",
        ]
    )
    writer.writerow(
        [
            "Laptop",
            "SN-EXAMPLE-1001",
            "Dell",
            "Latitude 5440",
            "2026-01-15",
            "2026-01-15",
            "2029-01-14",
            "Default Vendor",
            "85000",
            "HQ",
            "IT",
            "InStock",
        ]
    )
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=asset_bulk_upload_template.csv"},
    )


@app.get("/assets/search", response_model=List[AssetOut])
def search_assets(
    q: Optional[str] = None,
    asset_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    department: Optional[str] = None,
    location: Optional[str] = None,
    assignee_name: Optional[str] = None,
    assignee_email: Optional[str] = None,
    assignee_phone: Optional[str] = None,
    assignee_emp_id: Optional[str] = None,
    warranty_to: Optional[date] = None,
    assigned: Optional[bool] = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    has_filter = any([q, asset_type, status_filter, department, location, assignee_name, assignee_email, assignee_phone, assignee_emp_id, warranty_to, assigned is not None])
    if not has_filter:
        return []

    query = db.query(Asset).outerjoin(AssetHolder, AssetHolder.asset_id == Asset.id)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Asset.asset_id.ilike(like),
                Asset.serial_number.ilike(like),
                Asset.manufacturer.ilike(like),
                Asset.model.ilike(like),
                AssetHolder.user_name.ilike(like),
                AssetHolder.email.ilike(like),
                AssetHolder.phone.ilike(like),
                AssetHolder.emp_id.ilike(like),
            )
        )
    if asset_type:
        query = query.filter(Asset.asset_type == asset_type)
    if status_filter:
        query = query.filter(Asset.status == status_filter)
    if department:
        query = query.filter(Asset.department == department)
    if location:
        query = query.filter(Asset.location == location)
    if assignee_name:
        query = query.filter(AssetHolder.user_name.ilike(f"%{assignee_name}%"))
    if assignee_email:
        query = query.filter(AssetHolder.email.ilike(f"%{assignee_email}%"))
    if assignee_phone:
        query = query.filter(AssetHolder.phone.ilike(f"%{assignee_phone}%"))
    if assignee_emp_id:
        query = query.filter(AssetHolder.emp_id.ilike(f"%{assignee_emp_id}%"))
    if warranty_to:
        query = query.filter(Asset.warranty_end <= warranty_to)
    if assigned is not None:
        if assigned:
            query = query.filter(Asset.status == "Assigned")
        else:
            query = query.filter(Asset.status != "Assigned")
    return query.order_by(Asset.id.desc()).all()


@app.get("/assets/{asset_id}", response_model=AssetOut)
def get_asset(asset_id: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    asset = db.query(Asset).filter(Asset.asset_id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    return asset


@app.patch("/assets/{asset_id}", response_model=AssetOut)
def update_asset(asset_id: str, payload: AssetUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    asset = db.query(Asset).filter(Asset.asset_id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    data = payload.model_dump(exclude_unset=True)
    if current_user.role not in ["Admin", "ITUser"]:
        raise HTTPException(status_code=403, detail="Admin/IT access required")
    if current_user.role != "Admin":
        disallowed = [k for k in data.keys() if k != "location"]
        if disallowed:
            raise HTTPException(
                status_code=403,
                detail="Only Admin can update asset details. ITUser can only update location.",
            )
        if "location" not in data:
            raise HTTPException(status_code=400, detail="No updatable field provided")
    if "status" in data:
        require_active_status(data["status"], db)
    for key, value in data.items():
        setattr(asset, key, value)
    db.commit()
    db.refresh(asset)
    log_audit(db, action="ASSET_UPDATED", entity_type="Asset", entity_id=str(asset.id), actor_user_id=current_user.id, details=f"fields={','.join(data.keys())}")
    return asset


@app.get("/assets/{asset_id}/barcode")
def get_asset_barcode(asset_id: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    asset = db.query(Asset).filter(Asset.asset_id == asset_id).first()
    if not asset or not asset.barcode_path:
        raise HTTPException(status_code=404, detail="Barcode not found")
    return FileResponse(asset.barcode_path)


@app.get("/assets/{asset_id}/qr")
def get_asset_qr(asset_id: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    asset = db.query(Asset).filter(Asset.asset_id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    qr_path = asset_qr_path(asset.asset_id)
    if not os.path.exists(qr_path):
        generate_qr_label(asset)
    return FileResponse(qr_path)


@app.post("/assignments/assign", response_model=AssignmentOut)
def assign_asset(payload: AssignmentCreate, current_user: User = Depends(require_it_or_admin), db: Session = Depends(get_db)):
    asset = db.query(Asset).filter(Asset.id == payload.asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if asset.status == "Assigned":
        raise HTTPException(status_code=400, detail="Asset already assigned")
    if not payload.assignee_name or not payload.assignee_name.strip():
        raise HTTPException(status_code=400, detail="assignee_name is required for assignment")
    if not payload.department or not payload.department.strip():
        raise HTTPException(status_code=400, detail="department is required for assignment")
    if not payload.location or not payload.location.strip():
        raise HTTPException(status_code=400, detail="location is required for assignment")
    if not payload.remarks or not payload.remarks.strip():
        raise HTTPException(status_code=400, detail="remarks is required for assignment")

    new_assignment = Assignment(
        asset_id=payload.asset_id,
        user_id=payload.user_id,
        department=payload.department,
        location=payload.location,
        assign_date=datetime.utcnow().date(),
        expected_return=payload.expected_return,
        remarks=payload.remarks,
        assigned_by=current_user.id,
    )
    db.add(new_assignment)
    asset.status = "Assigned"
    upsert_asset_holder(
        db,
        asset.id,
        payload.assignee_name,
        payload.assignee_email,
        payload.assignee_phone,
        payload.assignee_emp_id,
        is_active=True,
    )

    ev_utc, ev_local = now_utc_local_pair()
    event = AssignmentEvent(
        asset_id=asset.id,
        event_type="ASSIGN",
        from_user_id=None,
        to_user_id=payload.user_id,
        from_department=None,
        to_department=payload.department,
        from_location=None,
        to_location=payload.location,
        expected_return=payload.expected_return,
        remarks=f"{payload.remarks} | assignee={payload.assignee_name or ''}, email={payload.assignee_email or ''}, phone={payload.assignee_phone or ''}, emp={payload.assignee_emp_id or ''}",
        created_by=current_user.id,
        created_at=ev_utc,
        created_at_utc=ev_utc,
        created_at_local=ev_local,
    )
    db.add(event)
    db.commit()
    db.refresh(new_assignment)

    log_audit(db, action="ASSET_ASSIGNED", entity_type="Asset", entity_id=str(asset.id), actor_user_id=current_user.id)
    return new_assignment


@app.post("/assignments/{asset_id}/return")
def return_asset(asset_id: int, remarks: str = Form(...), current_user: User = Depends(require_it_or_admin), db: Session = Depends(get_db)):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if asset.status != "Assigned":
        raise HTTPException(status_code=400, detail="Asset is not assigned")

    active_assignment = db.query(Assignment).filter(Assignment.asset_id == asset_id).order_by(Assignment.id.desc()).first()
    if not active_assignment:
        raise HTTPException(status_code=400, detail="No assignment record found")
    if not remarks or not remarks.strip():
        raise HTTPException(status_code=400, detail="remarks is required for return")

    ev_utc, ev_local = now_utc_local_pair()
    event = AssignmentEvent(
        asset_id=asset.id,
        event_type="RETURN",
        from_user_id=active_assignment.user_id,
        to_user_id=None,
        from_department=active_assignment.department,
        to_department=None,
        from_location=active_assignment.location,
        to_location=None,
        expected_return=None,
        remarks=remarks,
        created_by=current_user.id,
        created_at=ev_utc,
        created_at_utc=ev_utc,
        created_at_local=ev_local,
    )
    db.add(event)
    asset.status = "InStock"
    upsert_asset_holder(db, asset.id, None, None, None, None, is_active=False)
    db.commit()

    log_audit(db, action="ASSET_RETURNED", entity_type="Asset", entity_id=str(asset.id), actor_user_id=current_user.id)
    return {"message": "Asset returned successfully"}


@app.post("/assignments/{asset_id}/repair")
def move_asset_to_repair(
    asset_id: int,
    remarks: str = Form(...),
    current_user: User = Depends(require_it_or_admin),
    db: Session = Depends(get_db),
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    if asset.status == "Retired":
        raise HTTPException(status_code=400, detail="Retired asset cannot be moved to repair")
    if not remarks or not remarks.strip():
        raise HTTPException(status_code=400, detail="remarks is required for repair")
    ev_utc, ev_local = now_utc_local_pair()
    event = AssignmentEvent(
        asset_id=asset.id,
        event_type="REPAIR",
        from_user_id=None,
        to_user_id=None,
        from_department=asset.department,
        to_department=asset.department,
        from_location=asset.location,
        to_location=asset.location,
        expected_return=None,
        remarks=remarks,
        created_by=current_user.id,
        created_at=ev_utc,
        created_at_utc=ev_utc,
        created_at_local=ev_local,
    )
    db.add(event)
    asset.status = "UnderRepair"
    upsert_asset_holder(db, asset.id, None, None, None, None, is_active=False)
    db.commit()
    log_audit(db, action="ASSET_UNDER_REPAIR", entity_type="Asset", entity_id=str(asset.id), actor_user_id=current_user.id, details=remarks)
    return {"message": "Asset moved to repair"}


@app.get("/assets/{asset_id}/history", response_model=List[AssignmentEventOut])
def get_asset_history(asset_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    events = db.query(AssignmentEvent).filter(AssignmentEvent.asset_id == asset_id).order_by(AssignmentEvent.created_at.desc()).all()
    return events


@app.get("/assets/{asset_code}/timeline", response_model=List[TimelineEventOut])
def get_asset_timeline(asset_code: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    asset = db.query(Asset).filter(Asset.asset_id == asset_code).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    timeline: List[dict] = []
    timeline.append(
        {
            "timestamp": datetime.combine(asset.purchase_date, datetime.min.time()).replace(tzinfo=IST),
            "event_type": "ASSET_PURCHASED",
            "details": f"Asset purchased on {asset.purchase_date.isoformat()}",
        }
    )

    events = db.query(AssignmentEvent).filter(AssignmentEvent.asset_id == asset.id).all()
    audits = (
        db.query(AuditLog)
        .filter(AuditLog.entity_type == "Asset", AuditLog.entity_id == str(asset.id))
        .all()
    )

    actor_ids = set()
    for ev in events:
        if ev.created_by:
            actor_ids.add(ev.created_by)
    for log in audits:
        if log.actor_user_id:
            actor_ids.add(log.actor_user_id)

    user_map = {}
    if actor_ids:
        for u in db.query(User).filter(User.id.in_(list(actor_ids))).all():
            user_map[u.id] = u.username

    for ev in events:
        actor_name = f" by {user_map[ev.created_by]}" if ev.created_by in user_map else ""
        ts = as_ist_aware(ev.created_at_local, ev.created_at_utc or ev.created_at)
        timeline.append(
            {
                "timestamp": ts,
                "event_type": ev.event_type,
                "details": f"{ev.remarks}{actor_name}",
            }
        )
    for log in audits:
        actor = user_map.get(log.actor_user_id)
        actor_suffix = f" by {actor}" if actor else ""
        ts = as_ist_aware(log.created_at_local, log.created_at_utc or log.created_at)
        timeline.append(
            {
                "timestamp": ts,
                "event_type": log.action,
                "details": f"{(log.details or '').strip()}{actor_suffix}",
            }
        )

    timeline.sort(key=lambda x: x["timestamp"], reverse=True)
    return timeline


@app.get("/assignees/{user_name}", response_model=AssigneeSummaryOut)
def get_assignee_summary(user_name: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    normalized = user_name.strip()
    if not normalized:
        raise HTTPException(status_code=400, detail="Invalid assignee name")

    holders = (
        db.query(AssetHolder)
        .filter(AssetHolder.user_name == normalized)
        .order_by(AssetHolder.updated_at.desc())
        .all()
    )
    latest = holders[0] if holders else None

    active_holders = [h for h in holders if h.is_active]
    active_holder_map = {h.asset_id: h for h in active_holders}
    current_asset_ids = [h.asset_id for h in active_holders]

    all_events = db.query(AssignmentEvent).order_by(AssignmentEvent.created_at.asc()).all()
    marker = f"assignee={normalized}"
    assign_events = [ev for ev in all_events if ev.event_type == "ASSIGN" and ev.remarks and marker in ev.remarks]

    asset_ids = set([ev.asset_id for ev in assign_events] + current_asset_ids)
    assets_map = {}
    if asset_ids:
        for a in db.query(Asset).filter(Asset.id.in_(list(asset_ids))).all():
            assets_map[a.id] = a

    def asset_row(a: Asset):
        return {
            "id": a.id,
            "asset_id": a.asset_id,
            "asset_type": a.asset_type,
            "serial_number": a.serial_number,
            "manufacturer": a.manufacturer,
            "model": a.model,
            "purchase_date": a.purchase_date,
            "warranty_start": a.warranty_start,
            "warranty_end": a.warranty_end,
            "vendor": a.vendor,
            "cost": a.cost,
            "location": a.location,
            "department": a.department,
            "status": a.status,
            "barcode_path": a.barcode_path,
        }

    terminal_events = {"RETURN", "LOST", "SCRAP", "END_OF_LIFE"}
    current_assets = []
    historical_assets = []
    for ev in assign_events:
        asset = assets_map.get(ev.asset_id)
        if not asset:
            continue
        close_event = None
        for follow in all_events:
            follow_ts = follow.created_at_utc or follow.created_at
            ev_ts = ev.created_at_utc or ev.created_at
            if follow.asset_id == ev.asset_id and follow_ts > ev_ts and follow.event_type in terminal_events:
                close_event = follow
                break
        ev_local = as_ist_aware(ev.created_at_local, ev.created_at_utc or ev.created_at)
        close_local = as_ist_aware(
            close_event.created_at_local if close_event else None,
            (close_event.created_at_utc or close_event.created_at) if close_event else None,
        )
        row = {
            "asset": asset_row(asset),
            "assigned_at": ev_local,
            "closed_at": close_local,
            "closed_event": close_event.event_type if close_event else None,
        }
        if ev.asset_id in active_holder_map and not close_event:
            current_assets.append(row)
        elif close_event:
            historical_assets.append(row)

    return {
        "user_name": normalized,
        "email": latest.email if latest else None,
        "phone": latest.phone if latest else None,
        "emp_id": latest.emp_id if latest else None,
        "current_assets": current_assets,
        "historical_assets": historical_assets,
    }


@app.post("/assets/{asset_id}/lifecycle")
def update_asset_lifecycle(
    asset_id: int,
    payload: AssetLifecycleIn,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    asset = db.query(Asset).filter(Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    action_map = {
        "scrap": "Scrapped",
        "end_of_life": "EndOfLife",
        "lost": "Lost",
    }
    new_status = action_map[payload.action]
    if not payload.remarks or not payload.remarks.strip():
        raise HTTPException(status_code=400, detail="remarks is required for lifecycle actions")
    asset.status = new_status

    ev_utc, ev_local = now_utc_local_pair()
    event = AssignmentEvent(
        asset_id=asset.id,
        event_type=payload.action.upper(),
        from_user_id=None,
        to_user_id=None,
        from_department=asset.department,
        to_department=asset.department,
        from_location=asset.location,
        to_location=asset.location,
        expected_return=None,
        remarks=payload.remarks or f"Marked as {new_status}",
        created_by=current_user.id,
        created_at=ev_utc,
        created_at_utc=ev_utc,
        created_at_local=ev_local,
    )
    db.add(event)
    upsert_asset_holder(db, asset.id, None, None, None, None, is_active=False)
    db.commit()

    log_audit(
        db,
        action=f"ASSET_{payload.action.upper()}",
        entity_type="Asset",
        entity_id=str(asset.id),
        actor_user_id=current_user.id,
        details=payload.remarks,
    )
    return {"message": f"Asset marked as {new_status}"}


@app.get("/dashboard/summary", response_model=DashboardSummaryOut)
def dashboard_summary(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    today = datetime.utcnow().date()
    in_30_days = today + timedelta(days=30)

    total_assets = db.query(func.count(Asset.id)).scalar() or 0
    assigned_assets = db.query(func.count(Asset.id)).filter(Asset.status == "Assigned").scalar() or 0
    under_repair_assets = db.query(func.count(Asset.id)).filter(Asset.status == "UnderRepair").scalar() or 0
    unassigned_assets = total_assets - assigned_assets
    warranty_expiring = db.query(func.count(Asset.id)).filter(Asset.warranty_end >= today, Asset.warranty_end <= in_30_days).scalar() or 0

    by_type_rows = db.query(Asset.asset_type, func.count(Asset.id)).group_by(Asset.asset_type).order_by(func.count(Asset.id).desc()).all()
    by_department_rows = db.query(Asset.department, func.count(Asset.id)).group_by(Asset.department).order_by(func.count(Asset.id).desc()).all()

    return {
        "total_assets": total_assets,
        "assigned_assets": assigned_assets,
        "unassigned_assets": unassigned_assets,
        "under_repair_assets": under_repair_assets,
        "warranty_expiring_30_days": warranty_expiring,
        "by_type": [{"asset_type": x[0], "count": x[1]} for x in by_type_rows],
        "by_department": [{"department": x[0], "count": x[1]} for x in by_department_rows],
    }


@app.get("/alerts/warranty", response_model=List[WarrantyAlertOut])
def warranty_alerts(within_days: int = 30, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    today = datetime.utcnow().date()
    target = today + timedelta(days=within_days)
    rows = db.query(Asset).filter(Asset.warranty_end >= today, Asset.warranty_end <= target).order_by(Asset.warranty_end.asc()).all()
    result = []
    for asset in rows:
        result.append({
            "asset_id": asset.asset_id,
            "serial_number": asset.serial_number,
            "warranty_end": asset.warranty_end,
            "days_left": (asset.warranty_end - today).days,
        })
    return result


@app.get("/reports/assets.csv")
def export_assets_csv(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    rows = (
        db.query(Asset, AssetHolder)
        .outerjoin(AssetHolder, AssetHolder.asset_id == Asset.id)
        .order_by(Asset.id.asc())
        .all()
    )
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "id",
            "asset_id",
            "asset_type",
            "serial_number",
            "manufacturer",
            "model",
            "purchase_date",
            "warranty_start",
            "warranty_end",
            "vendor",
            "cost",
            "location",
            "department",
            "status",
            "assignee_name",
            "assignee_email",
            "assignee_phone",
            "assignee_emp_id",
            "assignment_state",
        ]
    )
    for a, h in rows:
        writer.writerow(
            [
                a.id,
                a.asset_id,
                a.asset_type,
                a.serial_number,
                a.manufacturer,
                a.model,
                a.purchase_date,
                a.warranty_start,
                a.warranty_end,
                a.vendor,
                a.cost,
                a.location,
                a.department,
                a.status,
                h.user_name if h else None,
                h.email if h else None,
                h.phone if h else None,
                h.emp_id if h else None,
                "Active" if (h and h.is_active) else "Inactive/None",
            ]
        )
    buffer.seek(0)
    return StreamingResponse(iter([buffer.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=assets_report.csv"})


@app.get("/reports/assignment-history.csv")
def export_assignment_history_csv(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    rows = db.query(AssignmentEvent).order_by(AssignmentEvent.created_at.desc()).all()
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["id", "asset_id", "event_type", "from_user_id", "to_user_id", "from_department", "to_department", "from_location", "to_location", "expected_return", "remarks", "created_by", "created_at_local"])
    for e in rows:
        ts_local = e.created_at_local or ist_from_utc_naive(e.created_at_utc or e.created_at)
        writer.writerow([e.id, e.asset_id, e.event_type, e.from_user_id, e.to_user_id, e.from_department, e.to_department, e.from_location, e.to_location, e.expected_return, e.remarks, e.created_by, ts_local])
    buffer.seek(0)
    return StreamingResponse(iter([buffer.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=assignment_history.csv"})


@app.get("/reports/current-assignments.csv")
def export_current_assignments_csv(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    rows = (
        db.query(Asset, AssetHolder)
        .join(AssetHolder, AssetHolder.asset_id == Asset.id)
        .filter(AssetHolder.is_active == True, Asset.status == "Assigned")
        .order_by(Asset.department.asc(), Asset.asset_id.asc())
        .all()
    )
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "asset_id",
            "asset_type",
            "serial_number",
            "department",
            "location",
            "assignee_name",
            "assignee_email",
            "assignee_phone",
            "assignee_emp_id",
            "assigned_or_updated_local",
        ]
    )
    for asset, holder in rows:
        assigned_local = holder.updated_at_local or ist_from_utc_naive(holder.updated_at_utc or holder.updated_at)
        writer.writerow(
            [
                asset.asset_id,
                asset.asset_type,
                asset.serial_number,
                asset.department,
                asset.location,
                holder.user_name,
                holder.email,
                holder.phone,
                holder.emp_id,
                assigned_local,
            ]
        )
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=current_assignments_report.csv"},
    )


@app.get("/reports/assets-by-status.csv")
def export_assets_by_status_csv(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    rows = (
        db.query(Asset.status, func.count(Asset.id).label("count"))
        .group_by(Asset.status)
        .order_by(func.count(Asset.id).desc())
        .all()
    )
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["status", "asset_count"])
    for status_name, count in rows:
        writer.writerow([status_name, count])
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=assets_by_status_report.csv"},
    )


@app.get("/reports/assets-by-department.csv")
def export_assets_by_department_csv(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    rows = (
        db.query(Asset.department, func.count(Asset.id).label("count"))
        .group_by(Asset.department)
        .order_by(func.count(Asset.id).desc())
        .all()
    )
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["department", "asset_count"])
    for dept, count in rows:
        writer.writerow([dept, count])
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=assets_by_department_report.csv"},
    )


@app.get("/reports/warranty-expiry.csv")
def export_warranty_expiry_csv(within_days: int = 60, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    today = datetime.utcnow().date()
    target = today + timedelta(days=within_days)
    rows = (
        db.query(Asset)
        .filter(Asset.warranty_end >= today, Asset.warranty_end <= target)
        .order_by(Asset.warranty_end.asc(), Asset.asset_id.asc())
        .all()
    )
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["asset_id", "asset_type", "serial_number", "department", "location", "warranty_end", "days_left", "status"])
    for a in rows:
        writer.writerow([a.asset_id, a.asset_type, a.serial_number, a.department, a.location, a.warranty_end, (a.warranty_end - today).days, a.status])
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=warranty_expiry_{within_days}d_report.csv"},
    )


@app.get("/reports/lifecycle-events.csv")
def export_lifecycle_events_csv(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    events = (
        db.query(AssignmentEvent, Asset)
        .join(Asset, Asset.id == AssignmentEvent.asset_id)
        .order_by(AssignmentEvent.created_at.desc())
        .all()
    )
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["asset_id", "event_type", "department", "location", "remarks", "created_by", "event_time_local"])
    for ev, asset in events:
        ts_local = ev.created_at_local or ist_from_utc_naive(ev.created_at_utc or ev.created_at)
        writer.writerow([asset.asset_id, ev.event_type, ev.to_department or ev.from_department, ev.to_location or ev.from_location, ev.remarks, ev.created_by, ts_local])
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=lifecycle_events_report.csv"},
    )


@app.get("/audit/logs", response_model=List[AuditLogOut])
def get_audit_logs(limit: int = 100, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    rows = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(limit).all()
    log_audit(db, action="AUDIT_VIEWED", entity_type="AuditLog", entity_id="bulk", actor_user_id=current_user.id, details=f"limit={limit}")
    return rows


@app.get("/branding", response_model=BrandingOut)
def get_branding():
    logo_path = get_logo_file()
    logo_url = None
    if logo_path and os.path.exists(logo_path):
        logo_url = f"/{logo_path}?v={int(os.path.getmtime(logo_path))}"
    return {"app_name": APP_NAME, "logo_url": logo_url}


@app.post("/branding/logo", response_model=BrandingOut)
async def upload_branding_logo(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    content_type = (file.content_type or "").lower()
    allowed_content_types = {"image/png": "png", "image/jpeg": "jpg", "image/webp": "webp", "image/gif": "gif"}
    ext = allowed_content_types.get(content_type)
    if not ext:
        # Fallback by filename extension if content type is generic/missing.
        name = (file.filename or "").lower()
        if name.endswith(".png"):
            ext = "png"
        elif name.endswith(".jpg") or name.endswith(".jpeg"):
            ext = "jpg"
        elif name.endswith(".webp"):
            ext = "webp"
        elif name.endswith(".gif"):
            ext = "gif"
    if not ext:
        raise HTTPException(status_code=400, detail="Only PNG, JPG, WEBP, GIF logo files are supported")

    # Remove previous logo files.
    for existing in os.listdir("static/branding"):
        if existing.lower().startswith("company_logo."):
            try:
                os.remove(os.path.join("static/branding", existing))
            except Exception:
                pass

    save_path = os.path.join("static/branding", f"company_logo.{ext}")
    with open(save_path, "wb") as out:
        shutil.copyfileobj(file.file, out)

    log_audit(
        db,
        action="BRANDING_LOGO_UPDATED",
        entity_type="Branding",
        entity_id="company_logo",
        actor_user_id=current_user.id,
        details=f"file={os.path.basename(save_path)}",
    )
    return get_branding()


@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})
